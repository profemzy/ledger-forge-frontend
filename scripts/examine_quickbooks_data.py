#!/usr/bin/env python3
"""
Examine QuickBooks Excel files and convert them to CSV for import
"""

import sys
import os
import openpyxl
import csv

def examine_excel_file(filepath):
    """Examine the structure of an Excel file"""
    print(f"\n{'='*80}")
    print(f"File: {os.path.basename(filepath)}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {ws.dimensions}")

        # Get headers (first row)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        print(f"Headers: {headers}")

        # Count rows with data
        row_count = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
        print(f"Data rows: {row_count}")

        # Show first 3 data rows
        print("\nSample data (first 3 rows):")
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 1):
            print(f"  Row {idx}: {row}")

    wb.close()

def convert_excel_to_csv(excel_path, output_dir):
    """Convert Excel file to CSV"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_filename = f"{base_name}_{sheet_name}.csv" if len(wb.sheetnames) > 1 else f"{base_name}.csv"
        csv_path = os.path.join(output_dir, csv_filename)

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None and str(cell).strip() for cell in row):
                    writer.writerow(row)

        print(f"✓ Converted {sheet_name} → {csv_filename}")

    wb.close()

def main():
    data_dir = "/Users/profemzy/RustroverProjects/ledger-forge/data"
    csv_dir = os.path.join(data_dir, "csv")

    # Create CSV directory if it doesn't exist
    os.makedirs(csv_dir, exist_ok=True)

    # Excel files to process
    excel_files = [
        "Trial_balance.xlsx",
        "Customers.xlsx",
        "Suppliers.xlsx",
        "Employees.xlsx",
        "Journal.xlsx",
        "General_ledger.xlsx",
        "Profit_and_loss.xlsx",
        "Balance_sheet.xlsx",
    ]

    print("QuickBooks Data Examination and Conversion")
    print("=" * 80)

    # Examine each file
    for filename in excel_files:
        filepath = os.path.join(data_dir, filename)
        if os.path.exists(filepath):
            examine_excel_file(filepath)
        else:
            print(f"\n⚠️  File not found: {filename}")

    # Convert to CSV
    print("\n" + "=" * 80)
    print("Converting to CSV...")
    print("=" * 80)

    for filename in excel_files:
        filepath = os.path.join(data_dir, filename)
        if os.path.exists(filepath):
            try:
                convert_excel_to_csv(filepath, csv_dir)
            except Exception as e:
                print(f"✗ Error converting {filename}: {e}")
        else:
            print(f"⚠️  Skipped {filename} (not found)")

    print(f"\n✅ CSV files saved to: {csv_dir}")

if __name__ == "__main__":
    main()
